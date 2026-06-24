from __future__ import annotations
import io
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from apps.investments.models import InvestmentProduct

HEADER_FILL = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
ALT_FILL = PatternFill(start_color='EBF3FC', end_color='EBF3FC', fill_type='solid')
INR_FORMAT = '₹#,##0.00'
PCT_FORMAT = '0.00%'


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)


def generate_excel_report(member=None, group=None, report_name='portfolio_snapshot') -> tuple[bytes, str]:
    owner_name = (member.name if member else group.name) if (member or group) else 'All'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_summary = wb.create_sheet('Summary')
    ws_summary.append(['Family Portfolio Intelligence'])
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.append([f'Report for: {owner_name}'])
    ws_summary.append([f'Generated: {date.today().strftime("%d %B %Y")}'])
    ws_summary.append([])

    if member:
        products = InvestmentProduct.objects.filter(
            investment_account__family_member=member, is_active=True
        )
    elif group:
        products = InvestmentProduct.objects.filter(
            investment_account__family_member__in=group.members.all(), is_active=True
        )
    else:
        products = InvestmentProduct.objects.none()

    total_invested = sum(float(p.invested_value) for p in products)
    total_current = sum(float(p.current_value) for p in products)
    gain = total_current - total_invested

    summary_rows = [
        ['Metric', 'Value'],
        ['Total Invested', total_invested],
        ['Current Value', total_current],
        ['Gain / Loss', gain],
        ['Return %', (gain / total_invested * 100) if total_invested else 0],
        ['Products Count', products.count()],
    ]
    for row in summary_rows:
        ws_summary.append(row)

    ws_summary['A5'].font = HEADER_FONT
    ws_summary['A5'].fill = HEADER_FILL

    # ── Sheet 2: Holdings ─────────────────────────────────────────────────────
    ws_holdings = wb.create_sheet('Holdings')
    headers = ['Name', 'ISIN', 'Type', 'Account', 'Invested (₹)', 'Current (₹)', 'Gain/Loss (₹)', 'G/L %', 'XIRR %', 'As of Date']
    ws_holdings.append(headers)

    for cell in ws_holdings[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    for i, p in enumerate(products.order_by('-current_value'), start=2):
        gl_pct = float(p.gain_loss_pct) / 100 if float(p.gain_loss_pct) else 0
        xirr = float(p.xirr) if p.xirr else None
        row = [
            p.name[:60],
            p.isin or '',
            p.product_type,
            p.investment_account.account_name,
            float(p.invested_value),
            float(p.current_value),
            float(p.gain_loss),
            gl_pct,
            xirr,
            str(p.as_of_date) if p.as_of_date else '',
        ]
        ws_holdings.append(row)
        if i % 2 == 0:
            for cell in ws_holdings[i]:
                cell.fill = ALT_FILL

        ws_holdings.cell(i, 5).number_format = INR_FORMAT
        ws_holdings.cell(i, 6).number_format = INR_FORMAT
        ws_holdings.cell(i, 7).number_format = INR_FORMAT
        ws_holdings.cell(i, 8).number_format = PCT_FORMAT
        if xirr is not None:
            ws_holdings.cell(i, 9).number_format = PCT_FORMAT

    _auto_width(ws_holdings)

    # ── Sheet 3: Asset Allocation ──────────────────────────────────────────────
    ws_alloc = wb.create_sheet('Asset Allocation')
    from collections import defaultdict
    type_vals = defaultdict(float)
    for p in products:
        type_vals[p.product_type] += float(p.current_value)

    ws_alloc.append(['Asset Type', 'Value (₹)', 'Allocation %'])
    for cell in ws_alloc[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for i, (t, v) in enumerate(sorted(type_vals.items(), key=lambda x: -x[1]), start=2):
        ws_alloc.append([t, v, v / total_current if total_current else 0])
        ws_alloc.cell(i, 2).number_format = INR_FORMAT
        ws_alloc.cell(i, 3).number_format = PCT_FORMAT

    ws_alloc.append(['TOTAL', total_current, 1.0])
    ws_alloc.cell(ws_alloc.max_row, 2).number_format = INR_FORMAT
    ws_alloc.cell(ws_alloc.max_row, 3).number_format = PCT_FORMAT
    _auto_width(ws_alloc)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    filename = f'portfolio_{owner_name.replace(" ", "_")}_{date.today().isoformat()}.xlsx'
    return excel_bytes, filename
