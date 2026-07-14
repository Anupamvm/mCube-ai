"""
Parser for MF Central CAS XLSX exports.

Sheet structure:
  "Portfolio Details"
    Rows 1-7: investor metadata (Name, Mobile, Email, PAN, From Date, To Date)
    Row 9-10: summary totals
    Row 12: headers — Scheme Name | AMC Name | Category | Folio No. | Invested Value | Current Value | Returns | Units
    Row 13+: one holding per row

  "Transaction Details"
    Rows 1-7: same investor metadata
    Row 9: headers — Scheme Name | Transaction Description | Date | NAV | Units | Amount
    Row 10+: one transaction per row

No ISIN column — holdings are keyed by (Folio No., Scheme Name).
"""

from __future__ import annotations
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation

logger = logging.getLogger('apps.investments')

_PORTFOLIO_SHEET  = 'Portfolio Details'
_TRANSACTION_SHEET = 'Transaction Details'
_DATE_FMTS = ('%d-%b-%Y', '%d-%B-%Y', '%Y-%m-%d', '%d/%m/%Y')


def _num(val) -> Decimal:
    try:
        return Decimal(str(val).replace(',', '').strip())
    except (InvalidOperation, TypeError):
        return Decimal('0')


def _parse_date(val):
    if val is None:
        return None
    s = str(val).strip()
    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _find_header_row(ws, required_cols: list[str]) -> int | None:
    """Return 0-based row index of the first row containing all required_cols (case-insensitive)."""
    required_lower = {c.lower() for c in required_cols}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = {str(c or '').strip().lower() for c in row}
        if required_lower.issubset(cells):
            return i
    return None


def _row_to_dict(header_row: tuple, data_row: tuple) -> dict:
    return {
        str(h or '').strip(): str(v if v is not None else '').strip()
        for h, v in zip(header_row, data_row)
    }


def parse_mfcentral_xlsx(raw_bytes: bytes) -> dict:
    """
    Parse an MF Central CAS XLSX file.

    Returns:
        {
          'investor': {pan, name, email, mobile},
          'holdings': [{scheme_name, amc, category, folio, invested_value, current_value, units, returns}],
          'transactions': [{scheme_name, description, date, nav, units, amount}],
          'summary': {total_invested, current_value, profit_loss},
        }
    Raises ValueError if the file doesn't look like an MF Central XLSX.
    """
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)

    if _PORTFOLIO_SHEET not in wb.sheetnames:
        raise ValueError(f'No "{_PORTFOLIO_SHEET}" sheet — not an MF Central CAS XLSX')

    ws_p = wb[_PORTFOLIO_SHEET]
    all_rows_p = list(ws_p.iter_rows(values_only=True))

    # ── Investor metadata ────────────────────────────────────────────────
    investor: dict[str, str] = {}
    for row in all_rows_p:
        vals = [str(v or '').strip() for v in row]
        key = vals[0].lower().rstrip(':')
        val = vals[1] if len(vals) > 1 else ''
        if key == 'name':
            investor['name'] = val
        elif key == 'pan':
            investor['pan'] = val
        elif key == 'email':
            investor['email'] = val
        elif 'mobile' in key:
            investor['mobile'] = val

    # ── Summary totals ────────────────────────────────────────────────────
    summary: dict[str, Decimal] = {}
    for i, row in enumerate(all_rows_p):
        labels = [str(v or '').strip().lower() for v in row]
        if 'total investments' in labels:
            data_row = all_rows_p[i + 1] if i + 1 < len(all_rows_p) else ()
            if data_row:
                label_row = row
                for lbl, val in zip(label_row, data_row):
                    k = str(lbl or '').strip().lower()
                    if k:
                        summary[k] = _num(val)
            break

    # ── Holdings ──────────────────────────────────────────────────────────
    hdridx = _find_header_row(ws_p, ['Scheme Name', 'Folio No.'])
    if hdridx is None:
        raise ValueError('Cannot find holdings header row in Portfolio Details sheet')

    header_row = all_rows_p[hdridx]
    holdings = []
    for row in all_rows_p[hdridx + 1:]:
        d = _row_to_dict(header_row, row)
        scheme = d.get('Scheme Name', '').strip()
        if not scheme:
            continue
        holdings.append({
            'scheme_name':    scheme,
            'amc':            d.get('AMC Name', '').strip(),
            'category':       d.get('Category', '').strip(),
            'folio':          d.get('Folio No.', '').strip(),
            'invested_value': _num(d.get('Invested Value', 0)),
            'current_value':  _num(d.get('Current Value', 0)),
            'returns':        _num(d.get('Returns', 0)),
            'units':          _num(d.get('Units', 0)),
        })

    # ── Transactions ──────────────────────────────────────────────────────
    transactions = []
    if _TRANSACTION_SHEET in wb.sheetnames:
        ws_t = wb[_TRANSACTION_SHEET]
        all_rows_t = list(ws_t.iter_rows(values_only=True))
        tidx = _find_header_row(ws_t, ['Scheme Name', 'Transaction Description'])
        if tidx is not None:
            theader = all_rows_t[tidx]
            for row in all_rows_t[tidx + 1:]:
                d = _row_to_dict(theader, row)
                scheme = d.get('Scheme Name', '').strip()
                if not scheme:
                    continue
                transactions.append({
                    'scheme_name': scheme,
                    'description': d.get('Transaction Description', '').strip(),
                    'date':        _parse_date(d.get('Date')),
                    'nav':         _num(d.get('NAV', 0)),
                    'units':       _num(d.get('Units', 0)),
                    'amount':      _num(d.get('Amount', 0)),
                })

    logger.info(
        'MF Central XLSX parsed: investor=%s  holdings=%d  txns=%d',
        investor.get('name', '?'), len(holdings), len(transactions),
    )
    return {
        'investor': investor,
        'holdings': holdings,
        'transactions': transactions,
        'summary': summary,
    }


def _txn_type(desc: str) -> str:
    d = desc.lower()
    if any(k in d for k in ('switch out', 'redeem')):
        return 'REDEMPTION'
    if 'switch in' in d:
        return 'SWITCH_IN'
    if 'dividend' in d:
        return 'DIVIDEND'
    return 'PURCHASE'


def save_mfcentral_xlsx(member, raw_bytes: bytes, filename: str, uploaded_by) -> dict:
    """
    Parse an MF Central XLSX CAS and upsert into the DB.
    Creates a CASUpload record and returns result stats.

    Dedup strategy:
    - XLSX has per-folio rows → aggregate by scheme name first (sum units + values).
    - Skip zero-unit / zero-value rows (closed positions).
    - Use resolve_mf_product to merge against any existing ISIN-keyed records.
    """
    from datetime import date as date_cls
    from collections import defaultdict
    from apps.investments.models import (
        CASUpload, InvestmentAccount, Transaction, UserCASPassword,
    )
    from apps.investments.services.product_resolver import resolve_mf_product
    from .base import mask_pan

    data = parse_mfcentral_xlsx(raw_bytes)
    investor = data['investor']
    pan = investor.get('pan', '')

    if pan:
        if not member.pan_masked:
            member.pan_masked = mask_pan(pan)
            member.save(update_fields=['pan_masked'])
        # Auto-learn: the CAS PDF password is always the PAN, so any future
        # password-protected upload for this member (e.g. NSDL e-CAS) unlocks
        # automatically via the existing saved_passwords mechanism.
        UserCASPassword.objects.get_or_create(
            user=uploaded_by, password=pan,
            defaults={'label': f'{member.name} PAN'},
        )

    account, _ = InvestmentAccount.objects.get_or_create(
        family_member=member,
        account_type='MF_FOLIO',
        data_source='CAS',
        defaults={'account_name': 'MF Central MF Folio'},
    )

    upload = CASUpload.objects.create(
        family_member=member,
        uploaded_by=uploaded_by,
        cas_type='MFCENTRAL',
        original_filename=filename,
        parse_status='PROCESSING',
    )

    # ── Aggregate per-folio rows into per-scheme totals ───────────────────
    agg: dict[str, dict] = {}
    for h in data['holdings']:
        name = h['scheme_name']
        if not name:
            continue
        if name not in agg:
            agg[name] = {'invested': Decimal('0'), 'current': Decimal('0'),
                         'units': Decimal('0'), 'amc': h['amc'], 'category': h['category'], 'folios': []}
        e = agg[name]
        e['invested'] += h['invested_value']
        e['current']  += h['current_value']
        e['units']    += h['units']
        if h['category'] and not e['category']:
            e['category'] = h['category']
        folio = h['folio']
        if folio and folio not in e['folios']:
            e['folios'].append(folio)

    products_created = products_updated = transactions_created = 0

    for scheme_name, e in agg.items():
        # Skip closed / zero-value positions — mark existing as inactive if present
        if e['units'] == 0 and e['current'] == 0:
            from apps.investments.models import InvestmentProduct
            InvestmentProduct.objects.filter(
                investment_account=account,
                product_type='MUTUAL_FUND',
                name=scheme_name,
            ).update(is_active=False)
            continue

        defaults = {
            'data_source': 'CAS',
            'is_active': True,
            'invested_value': e['invested'].quantize(Decimal('0.01')),
            'current_value':  e['current'].quantize(Decimal('0.01')),
            'as_of_date': date_cls.today(),
            'extra_data': {'folios': e['folios'], 'amc': e['amc'],
                           'units': float(e['units']),
                           **({'category': e['category']} if e['category'] else {})},
            'source_upload': upload,
        }
        _product, created = resolve_mf_product(account, scheme_name, isin='', defaults=defaults)
        if created:
            products_created += 1
        else:
            products_updated += 1

    # ── Transactions ──────────────────────────────────────────────────────
    for t in data['transactions']:
        if not t['scheme_name'] or not t['date']:
            continue
        from apps.investments.models import InvestmentProduct
        product = InvestmentProduct.objects.filter(
            investment_account=account,
            name=t['scheme_name'],
        ).order_by('-isin').first()

        units = float(t['units'])
        amount = float(t['amount'])
        amount_decimal = Decimal(str(round(amount, 2)))
        description = (t['description'] or '')[:500]
        # order_no is always blank for MF Central data, so amount alone isn't
        # a safe differentiator with years of history — two distinct same-day
        # transactions of a round-number amount (e.g. two ₹5,000 SIP
        # instalments) are a real possibility. description is a raw cell
        # value straight from the sheet (not fuzzy-regenerated), so it's
        # stable across re-parses and narrows the key further.
        #
        # Match by `product` (resolved by name, stable) rather than
        # `product.isin` — the product's own ISIN can go from blank to
        # resolved between two uploads (via the MFAPI name-matcher backfill),
        # which would otherwise silently break this lookup and duplicate
        # every transaction the next time the same statement is re-uploaded.
        lookup = dict(
            investment_account=account,
            transaction_date=t['date'],
            order_no='',
            amount=amount_decimal,
            description=description,
        )
        if product:
            lookup['product'] = product
        else:
            lookup['isin'] = ''
            lookup['security_name'] = t['scheme_name']
        txn_obj, created = Transaction.objects.get_or_create(
            **lookup,
            defaults={
                'product': product,
                'isin': product.isin if product else '',
                'cas_upload': upload,
                'security_name': t['scheme_name'],
                'transaction_type': _txn_type(t['description']),
                'units_debit': max(0.0, -units),
                'units_credit': max(0.0, units),
                'units_balance': 0.0,
                'nav_at_transaction': t['nav'],
            },
        )
        if created:
            transactions_created += 1
        elif txn_obj.cas_upload_id != upload.id:
            # Re-attribute ownership to the most recent upload that
            # reconfirmed this transaction, mirroring product ownership.
            txn_obj.cas_upload = upload
            txn_obj.save(update_fields=['cas_upload'])

    upload.products_created = products_created
    upload.products_updated = products_updated
    upload.transactions_created = transactions_created
    upload.parse_status = 'DONE'
    upload.save()

    return {
        'products_created': products_created,
        'products_updated': products_updated,
        'transactions_created': transactions_created,
        'upload_id': upload.id,
    }
