"""
CSV Importers for Broker Trade History

Parses F&O trade history from CSV/Excel files:
- Kotak Neo: Gain/Loss CSV or Excel (Derivatives section)
- ICICI Breeze: FNO Portfolio CSV
"""

import csv
import io
import logging
import re
import uuid
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple

from django.core.files.uploadedfile import InMemoryUploadedFile, TemporaryUploadedFile
from django.db import transaction

from apps.brokers.models import BrokerContractPnL, CSVImportLog

logger = logging.getLogger(__name__)


# Symbol mapping for Breeze CSV (abbreviated to full symbol names)
BREEZE_SYMBOL_MAP = {
    'CNXBAN': 'BANKNIFTY',
    'HDFBAN': 'HDFCBANK',
    'NIFTY': 'NIFTY',
    'NIFTY50': 'NIFTY',
    'ICIBAN': 'ICICIBANK',
    'AXSBNK': 'AXISBANK',
    'SBIBNK': 'SBIN',
    'STABAN': 'SBIN',
    'KOTBNK': 'KOTAKBANK',
    'RELIND': 'RELIANCE',
    'INFTEC': 'INFY',
    'TATMOT': 'TATAMOTORS',
    'TATSTL': 'TATASTEEL',
    'HDFC': 'HDFC',
    'LT': 'LT',
    'TCS': 'TCS',
    'SUNPHA': 'SUNPHARMA',
    'TITIND': 'TITAN',
    'PERSYS': 'PERSISTENT',
    'BAJFI': 'BAJFINANCE',
    'JIOFIN': 'JIOFIN',
    'DIXTEC': 'DIXON',
    'ASIPAI': 'ASIANPAINT',
    'AVESUP': 'AUROPHARMA',
    'VARBEV': 'VBL',
}


def parse_decimal(value: str, default: Decimal = Decimal('0.00')) -> Decimal:
    """Parse a decimal value from string, handling commas and empty values."""
    if not value or value.strip() == '' or value.strip() == '-' or value.strip().upper() == 'NA':
        return default
    try:
        # Remove commas and parentheses (for negative values)
        cleaned = value.replace(',', '').replace('(', '-').replace(')', '').strip()
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        logger.warning(f"Could not parse decimal: {value}")
        return default


def determine_fy_from_date(d: date) -> str:
    """
    Determine Financial Year from a date.

    Indian FY runs April 1 to March 31.
    - If date is April or later, FY is current_year-next_year
    - If date is Jan-Mar, FY is previous_year-current_year

    Args:
        d: A date object

    Returns:
        FY string like '2024-25'
    """
    if not d:
        return ''

    year = d.year
    month = d.month

    if month >= 4:
        # April or later - we're in FY starting this year
        return f"{year}-{str(year + 1)[-2:]}"
    else:
        # Jan-Mar - we're in FY that started last year
        return f"{year - 1}-{str(year)[-2:]}"


def parse_int(value: str, default: int = 0) -> int:
    """Parse an integer value from string."""
    if not value or value.strip() == '' or value.strip() == '-' or value.strip().upper() == 'NA':
        return default
    try:
        cleaned = value.replace(',', '').strip()
        return int(float(cleaned))
    except (ValueError, TypeError):
        logger.warning(f"Could not parse int: {value}")
        return default


def convert_excel_to_csv_file(excel_file) -> InMemoryUploadedFile:
    """
    Convert an Excel (.xlsx/.xls) file to a CSV-like in-memory file
    that can be processed by existing CSV importers.

    Preserves the original filename pattern (replacing extension with .csv)
    so that FY extraction from filename still works.

    Args:
        excel_file: Django UploadedFile (InMemoryUploadedFile or TemporaryUploadedFile)

    Returns:
        InMemoryUploadedFile with CSV content, original filename with .csv extension
    """
    import openpyxl

    original_name = excel_file.name
    csv_name = re.sub(r'\.(xlsx?|xls)$', '.csv', original_name, flags=re.IGNORECASE)

    logger.info(f"Converting Excel file '{original_name}' to CSV format")

    # Read Excel file
    excel_file.seek(0)
    wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)
    ws = wb.active

    # Write rows to CSV buffer
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)

    for row in ws.iter_rows(values_only=True):
        # Convert None to empty string, numbers/dates to string
        csv_row = []
        for cell in row:
            if cell is None:
                csv_row.append('')
            else:
                csv_row.append(str(cell))
        writer.writerow(csv_row)

    wb.close()

    # Create InMemoryUploadedFile from CSV content
    csv_content = csv_buffer.getvalue().encode('utf-8-sig')
    csv_file = InMemoryUploadedFile(
        file=io.BytesIO(csv_content),
        field_name='csv_files',
        name=csv_name,
        content_type='text/csv',
        size=len(csv_content),
        charset='utf-8',
    )

    logger.info(f"Excel conversion complete: {ws.max_row} rows, {len(csv_content)} bytes CSV")
    return csv_file


class KotakFNOImporter:
    """
    Parse Kotak Neo Gain/Loss CSV - ALL sections (Equity, Mutual Funds, ETF, Derivatives).

    CSV Format (actual structure):
    - Rows 1-4: Client info header (Client Name, Client Code, Transaction Period, Transaction Type)
    - Row 5: Empty
    - Row 6: Column headers
    - Row 7+: Data sections with section markers:
        - "Equity" section -> Equity stocks
        - "Mutual Funds" section -> MF holdings
        - "ETF" section -> ETFs
        - "Derivatives" section -> F&O (Futures & Options)
    - Disclaimer at the end

    Column headers (Row 6):
    Script Name, Security Type, ISIN, Quantity, Buy Amt. (A), Sell Amt. (B), Intraday, <1yr., >1yr.,
    Total(T = B - A), GST (C), Brokerage (D), Misc. (E), STT/CTT (S), Total (C+D+E+S), Net P&L (T-S), Gross P&L (T+(C+D+E))

    Security Types:
    - EQUITY STOCK -> Equity segment
    - MUTUAL FUND -> Mutual Fund segment
    - ETF -> ETF segment
    - FUTSTK, FUTIDX -> Futures segment
    - OPTIDX, OPTSTK -> Options segment

    Contract formats (Derivatives):
      - Futures: "BANKNIFTY 25NOV25 XX 0" (SYMBOL DDMMMYY XX 0)
      - Options: "NIFTY 02DEC25 CE 26850" (SYMBOL DDMMMYY CE/PE STRIKE)

    Incremental Upload:
      - Only creates new records (not previously imported)
      - Only updates records where values have changed
      - Skips unchanged records to avoid unnecessary writes
      - Latest report takes priority when updating
    """

    # Section markers in the CSV
    SECTION_MARKERS = ['equity', 'mutual funds', 'etf', 'derivatives']

    # Map security types to segments
    SECURITY_TYPE_TO_SEGMENT = {
        'EQUITY STOCK': 'EQUITY',
        'MUTUAL FUND': 'MUTUAL_FUND',
        'ETF': 'ETF',
        'FUTSTK': 'FUTURES',
        'FUTIDX': 'FUTURES',
        'OPTIDX': 'OPTIONS',
        'OPTSTK': 'OPTIONS',
    }

    # Map security types for model storage (normalize names)
    SECURITY_TYPE_NORMALIZE = {
        'EQUITY STOCK': 'EQUITY_STOCK',
        'MUTUAL FUND': 'MUTUAL_FUND',
        'ETF': 'ETF',
        'FUTSTK': 'FUTSTK',
        'FUTIDX': 'FUTIDX',
        'OPTIDX': 'OPTIDX',
        'OPTSTK': 'OPTSTK',
    }

    # Fields to compare for detecting changes (excluding raw_data and import_batch_id)
    COMPARE_FIELDS = [
        'quantity', 'buy_amount', 'sell_amount', 'gross_pnl', 'net_pnl',
        'gst', 'brokerage', 'stt', 'misc_charges', 'total_charges'
    ]

    def __init__(self):
        self.batch_id = None
        self.errors = []
        self.records_created = 0
        self.records_updated = 0
        self.records_skipped = 0
        self.records_unchanged = 0  # New: track records with no changes
        self.section_counts = {}  # Track records per section
        self.fy = None  # Financial year extracted from filename

    def _parse_fy_from_filename(self, filename: str) -> str:
        """
        Extract Financial Year from Kotak filename.

        Filename format: Gain_Loss_A0YPQ_20240401_20250331.csv
                    or:  Gain_Loss_A0YPQ_20240401_20250331 (1).csv (with download counter)
        The dates represent the FY period (YYYYMMDD_YYYYMMDD).

        Returns:
            FY string like '2024-25' or empty string if cannot parse
        """
        try:
            # Extract date portion using regex - handle optional space and (N) before .csv
            match = re.search(r'_(\d{8})_(\d{8})(?:\s*\(\d+\))?\.', filename)
            if match:
                start_date_str = match.group(1)
                # Parse start date (YYYYMMDD format)
                start_year = int(start_date_str[:4])
                start_month = int(start_date_str[4:6])

                # FY starts in April
                if start_month >= 4:
                    # Start date is in April or later, this is the FY start year
                    fy = f"{start_year}-{str(start_year + 1)[-2:]}"
                else:
                    # Start date is Jan-Mar, this is the FY end year
                    fy = f"{start_year - 1}-{str(start_year)[-2:]}"

                logger.info(f"Extracted FY from filename {filename}: {fy}")
                return fy
        except Exception as e:
            logger.warning(f"Could not extract FY from filename {filename}: {e}")

        return ''

    def parse_contract(self, script_name: str, security_type: str) -> Dict:
        """
        Parse Kotak contract/script name to extract symbol, expiry, strike, option_type.

        For Equity/MF/ETF:
        - Symbol is the script name itself

        For Derivatives:
        - "BANKNIFTY 25NOV25 XX 0" -> Futures
        - "NIFTY 02DEC25 CE 26850" -> Call Option
        - "RELIANCE 26DEC25 PE 1200" -> Put Option
        """
        segment = self.SECURITY_TYPE_TO_SEGMENT.get(security_type, 'EQUITY')
        normalized_security_type = self.SECURITY_TYPE_NORMALIZE.get(security_type, security_type)

        result = {
            'symbol': script_name,
            'expiry_date': None,
            'strike_price': None,
            'option_type': '',
            'segment': segment,
            'security_type': normalized_security_type,
            'isin': '',
        }

        if not script_name:
            return result

        # For non-derivatives, symbol is the full script name
        if segment not in ['FUTURES', 'OPTIONS']:
            result['symbol'] = script_name
            return result

        # Parse derivative contracts
        parts = script_name.strip().split()
        if len(parts) < 2:
            result['symbol'] = script_name
            return result

        # First part is always the symbol
        result['symbol'] = parts[0]

        # Second part is the expiry date (DDMMMYY format)
        if len(parts) >= 2:
            try:
                expiry_str = parts[1].upper()
                # Handle formats: 25NOV25, 02DEC25
                result['expiry_date'] = datetime.strptime(expiry_str, '%d%b%y').date()
            except ValueError:
                logger.warning(f"Could not parse expiry date: {parts[1]}")

        # Determine if this is futures or options based on security_type
        if segment == 'OPTIONS':
            # Third part is CE/PE, fourth is strike
            if len(parts) >= 3:
                opt_type = parts[2].upper()
                if opt_type in ['CE', 'PE']:
                    result['option_type'] = opt_type
            if len(parts) >= 4:
                try:
                    result['strike_price'] = Decimal(parts[3])
                except (InvalidOperation, ValueError):
                    pass
        else:
            # Fallback: check if third part is CE/PE (in case security_type wasn't set)
            if len(parts) >= 3 and parts[2].upper() in ['CE', 'PE']:
                result['segment'] = 'OPTIONS'
                result['option_type'] = parts[2].upper()
                if len(parts) >= 4:
                    try:
                        result['strike_price'] = Decimal(parts[3])
                    except (InvalidOperation, ValueError):
                        pass
            elif len(parts) >= 3 and parts[2].upper() == 'XX':
                result['segment'] = 'FUTURES'

        return result

    def _find_data_start(self, rows: List[List[str]]) -> int:
        """
        Find where the actual data starts (after header rows).
        The column headers row contains 'Script Name' as first column.
        Returns the index of the first data row (after column headers).
        """
        for i, row in enumerate(rows):
            if not row or len(row) == 0:
                continue
            first_cell = str(row[0]).strip().lower()
            # Column headers row
            if first_cell == 'script name':
                logger.info(f"Found column headers at row {i+1}")
                return i + 1  # Data starts on next row

        # Fallback: assume data starts at row 7 (0-indexed row 6)
        logger.warning("Could not find 'Script Name' header, assuming data starts at row 7")
        return 6

    def import_file(self, file_obj, user=None) -> Dict:
        """
        Import ALL data from Kotak Neo Gain/Loss CSV file.

        Parses all sections: Equity, Mutual Funds, ETF, and Derivatives.

        Args:
            file_obj: File object or file path
            user: User performing the import

        Returns:
            dict with import results including per-section counts
        """
        self.batch_id = f"KOTAK_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        self.errors = []
        self.records_created = 0
        self.records_updated = 0
        self.records_skipped = 0
        self.records_unchanged = 0
        self.section_counts = {
            'equity': 0,
            'mutual_funds': 0,
            'etf': 0,
            'derivatives': 0,
        }

        # Get filename and extract FY
        if hasattr(file_obj, 'name'):
            filename = file_obj.name
        else:
            filename = 'kotak_gainloss.csv'

        # Extract FY from filename
        self.fy = self._parse_fy_from_filename(filename)

        # Create import log
        import_log = CSVImportLog.objects.create(
            import_batch_id=self.batch_id,
            file_type='KOTAK_FNO',
            original_filename=filename,
            status='PROCESSING',
            imported_by=user,
        )

        try:
            # Read CSV content
            if hasattr(file_obj, 'read'):
                content = file_obj.read()
                if isinstance(content, bytes):
                    content = content.decode('utf-8-sig')  # Handle BOM
            else:
                with open(file_obj, 'r', encoding='utf-8-sig') as f:
                    content = f.read()

            # Parse CSV
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)

            logger.info(f"Read {len(rows)} rows from Kotak Neo CSV")

            # Find where data starts
            data_start = self._find_data_start(rows)

            # Process all data rows
            current_section = None
            with transaction.atomic():
                for i in range(data_start, len(rows)):
                    row = rows[i]

                    # Skip empty rows
                    if not row or len(row) < 2:
                        continue

                    first_cell = str(row[0]).strip()
                    first_cell_lower = first_cell.lower()

                    # Check if this is a section marker
                    if first_cell_lower in self.SECTION_MARKERS:
                        current_section = first_cell_lower.replace(' ', '_')
                        logger.info(f"Entering section: {current_section} at row {i+1}")
                        continue

                    # Stop at disclaimer
                    if first_cell_lower == 'disclaimer':
                        logger.info(f"Reached disclaimer at row {i+1}, stopping")
                        break

                    # Skip empty script names or section markers with no data
                    if not first_cell:
                        continue

                    # Get security type from column 1
                    security_type = str(row[1]).strip() if len(row) > 1 else ''

                    # Skip if no valid security type
                    if not security_type or security_type not in self.SECURITY_TYPE_TO_SEGMENT:
                        # Could be a header row or invalid row
                        if first_cell_lower not in ['', 'script name']:
                            logger.debug(f"Skipping row {i+1}: unknown security type '{security_type}'")
                        self.records_skipped += 1
                        continue

                    try:
                        self._process_kotak_row(row, i + 1)
                        # Track section counts
                        segment = self.SECURITY_TYPE_TO_SEGMENT.get(security_type)
                        if segment == 'EQUITY':
                            self.section_counts['equity'] += 1
                        elif segment == 'MUTUAL_FUND':
                            self.section_counts['mutual_funds'] += 1
                        elif segment == 'ETF':
                            self.section_counts['etf'] += 1
                        elif segment in ['FUTURES', 'OPTIONS']:
                            self.section_counts['derivatives'] += 1
                    except Exception as e:
                        error_msg = f"Row {i+1}: {str(e)}"
                        self.errors.append(error_msg)
                        logger.warning(error_msg)

            # Update import log
            import_log.records_created = self.records_created
            import_log.records_updated = self.records_updated
            import_log.records_skipped = self.records_skipped
            import_log.errors = self.errors
            import_log.status = 'SUCCESS' if not self.errors else 'PARTIAL'
            import_log.save()

            total_records = self.records_created + self.records_updated
            logger.info(f"Import complete: {total_records} records ({self.records_created} created, {self.records_updated} updated, {self.records_unchanged} unchanged)")
            logger.info(f"Section breakdown: {self.section_counts}")

            return {
                'success': True,
                'batch_id': self.batch_id,
                'records_created': self.records_created,
                'records_updated': self.records_updated,
                'records_skipped': self.records_skipped,
                'records_unchanged': self.records_unchanged,
                'errors': self.errors,
                'section_counts': self.section_counts,
            }

        except Exception as e:
            logger.exception(f"Error importing Kotak CSV: {e}")
            import_log.status = 'FAILED'
            import_log.errors = [str(e)]
            import_log.save()

            return {
                'success': False,
                'batch_id': self.batch_id,
                'error': str(e),
                'errors': [str(e)],
            }

    def _process_kotak_row(self, row: List[str], row_num: int = 0):
        """
        Process a single row from Kotak CSV with incremental upload support.

        Incremental behavior:
        - Creates new records if they don't exist
        - Updates only if values have changed (latest report takes priority)
        - Skips unchanged records to avoid unnecessary database writes

        Column mapping (0-indexed):
        0: Script Name
        1: Security Type (EQUITY STOCK, MUTUAL FUND, ETF, FUTSTK, FUTIDX, OPTIDX, OPTSTK)
        2: ISIN
        3: Quantity
        4: Buy Amt. (A)
        5: Sell Amt. (B)
        6: Intraday
        7: <1yr. (Short-term P&L)
        8: >1yr. (Long-term P&L)
        9: Total(T = B - A) - Gross P&L
        10: GST (C)
        11: Brokerage (D)
        12: Misc. (E)
        13: STT/CTT (S)
        14: Total (C+D+E+S) - Total charges
        15: Net P&L (T-S)
        16: Gross P&L (T+(C+D+E))
        """
        script_name = str(row[0]).strip()
        security_type = str(row[1]).strip() if len(row) > 1 else ''
        isin = str(row[2]).strip() if len(row) > 2 else ''

        # Parse contract/script details
        contract = self.parse_contract(script_name, security_type)

        # Parse numeric values with correct column indices
        quantity = parse_int(row[3] if len(row) > 3 else '')
        buy_amount = parse_decimal(row[4] if len(row) > 4 else '')
        sell_amount = parse_decimal(row[5] if len(row) > 5 else '')

        # P&L values
        short_term_pnl = parse_decimal(row[7] if len(row) > 7 else '')  # <1yr
        long_term_pnl = parse_decimal(row[8] if len(row) > 8 else '')   # >1yr
        gross_pnl = parse_decimal(row[9] if len(row) > 9 else '')       # Total(T = B - A)

        # Charges
        gst = parse_decimal(row[10] if len(row) > 10 else '')
        brokerage = parse_decimal(row[11] if len(row) > 11 else '')
        misc_charges = parse_decimal(row[12] if len(row) > 12 else '')
        stt = parse_decimal(row[13] if len(row) > 13 else '')
        total_charges = parse_decimal(row[14] if len(row) > 14 else '')
        net_pnl = parse_decimal(row[15] if len(row) > 15 else '')

        # Store additional data in raw_data
        raw_data = {
            'row': row,
            'row_number': row_num,
            'isin': isin,
            'short_term_pnl': str(short_term_pnl),
            'long_term_pnl': str(long_term_pnl),
        }

        # Determine FY for this record
        # Use filename FY to match Kotak Neo UI exactly
        # Kotak Neo reports all records in a file under the file's FY period,
        # regardless of expiry date (e.g., TITAN 24APR25 is included in FY 2024-25 file totals)
        if self.fy:
            record_fy = self.fy
            logger.debug(f"Using filename FY {record_fy} for {script_name}")
        elif contract['expiry_date']:
            # Fallback: infer from expiry_date if filename FY not available
            record_fy = determine_fy_from_date(contract['expiry_date'])
            logger.debug(f"Inferred FY {record_fy} from expiry_date {contract['expiry_date']} for {script_name}")
        else:
            # Last resort: use current date FY
            record_fy = determine_fy_from_date(datetime.now().date())
            logger.debug(f"Using current date FY {record_fy} for {script_name}")

        # New values from the CSV (latest report)
        new_values = {
            'import_batch_id': self.batch_id,
            'symbol': contract['symbol'],
            'segment': contract['segment'],
            'security_type': contract['security_type'],
            'expiry_date': contract['expiry_date'],
            'strike_price': contract['strike_price'],
            'option_type': contract['option_type'],
            'quantity': quantity,
            'buy_amount': buy_amount,
            'sell_amount': sell_amount,
            'gross_pnl': gross_pnl,
            'net_pnl': net_pnl,
            'gst': gst,
            'brokerage': brokerage,
            'stt': stt,
            'misc_charges': misc_charges,
            'total_charges': total_charges,
            'raw_data': raw_data,
        }

        # Check if record already exists
        try:
            existing = BrokerContractPnL.objects.get(
                broker='KOTAK',
                trading_symbol=script_name,
                fy=record_fy
            )

            # Compare key fields to detect if anything has changed
            has_changes = False
            for field in self.COMPARE_FIELDS:
                old_value = getattr(existing, field)
                new_value = new_values.get(field)
                if old_value != new_value:
                    has_changes = True
                    logger.debug(f"Change detected in {script_name}.{field}: {old_value} -> {new_value}")
                    break

            if has_changes:
                # Update with new values (latest report takes priority)
                for key, value in new_values.items():
                    setattr(existing, key, value)
                existing.save()
                self.records_updated += 1
                logger.debug(f"Updated record for {script_name} ({contract['segment']}): Net P&L = {net_pnl}")
            else:
                # No changes detected, skip this record
                self.records_unchanged += 1
                logger.debug(f"Unchanged record for {script_name} ({contract['segment']}): Net P&L = {net_pnl}")

        except BrokerContractPnL.DoesNotExist:
            # Create new record
            BrokerContractPnL.objects.create(
                broker='KOTAK',
                trading_symbol=script_name,
                fy=record_fy,
                **new_values
            )
            self.records_created += 1
            logger.debug(f"Created record for {script_name} ({contract['segment']}): Net P&L = {net_pnl}, FY = {record_fy}")


class BreezeFNOImporter:
    """
    Parse Breeze FNO Portfolio CSV.

    CSV Format (actual structure):
    - Row 1: "Futures," section marker
    - Row 2: Headers - Contract,Trade Flow,Open Position Qty,...
    - Row 3+: Data rows

    Contract format: "FUT-HDFBAN-24-Feb-2026" or "OPT-NIFTY-24-Apr-2025-CE-22500"
    Contains: Realized P&L, Unrealized P&L per contract

    Incremental Upload:
      - Only creates new records (not previously imported)
      - Only updates records where values have changed
      - Skips unchanged records to avoid unnecessary writes
      - Latest report takes priority when updating
    """

    # Fields to compare for detecting changes (excluding raw_data and import_batch_id)
    COMPARE_FIELDS = [
        'quantity', 'buy_amount', 'gross_pnl', 'net_pnl',
        'realized_pnl', 'unrealized_pnl'
    ]

    def __init__(self):
        self.batch_id = None
        self.errors = []
        self.records_created = 0
        self.records_updated = 0
        self.records_skipped = 0
        self.records_unchanged = 0  # Track records with no changes
        self.fy = None  # Financial year

    def _determine_fy_from_current_date(self) -> str:
        """
        Determine Financial Year based on current date.

        Breeze FNO Portfolio CSV shows current portfolio status,
        so the realized P&L is for the current FY.

        Indian FY runs April 1 to March 31.
        - If current month >= April (4), FY is current_year-next_year
        - If current month < April, FY is previous_year-current_year

        Returns:
            FY string like '2024-25'
        """
        today = datetime.now()
        year = today.year
        month = today.month

        if month >= 4:
            # April or later - we're in FY starting this year
            fy = f"{year}-{str(year + 1)[-2:]}"
        else:
            # Jan-Mar - we're in FY that started last year
            fy = f"{year - 1}-{str(year)[-2:]}"

        logger.info(f"Determined FY from current date: {fy}")
        return fy

    def parse_contract(self, contract_str: str) -> Dict:
        """
        Parse Breeze contract string.

        Examples:
        - "FUT-HDFBAN-24-Feb-2026" -> HDFCBANK Futures Feb 2026
        - "FUT-CNXBAN-31-Jul-2025" -> BANKNIFTY Futures Jul 2025
        - "OPT-NIFTY-24-Apr-2025-CE-22500" -> NIFTY Call Option
        """
        result = {
            'symbol': '',
            'expiry_date': None,
            'strike_price': None,
            'option_type': '',
            'segment': 'FUTURES',
            'security_type': '',
        }

        if not contract_str:
            return result

        parts = contract_str.strip().split('-')
        if len(parts) < 4:
            result['symbol'] = contract_str
            return result

        # First part: FUT or OPT
        contract_type = parts[0].upper()
        if contract_type == 'FUT':
            result['segment'] = 'FUTURES'
            result['security_type'] = 'FUTIDX'  # Will be refined based on symbol
        elif contract_type == 'OPT':
            result['segment'] = 'OPTIONS'
            result['security_type'] = 'OPTIDX'  # Will be refined based on symbol

        # Second part: Symbol (may need mapping)
        raw_symbol = parts[1].upper()
        result['symbol'] = BREEZE_SYMBOL_MAP.get(raw_symbol, raw_symbol)

        # Refine security type based on symbol
        index_symbols = ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY', 'CNXBAN']
        if result['symbol'] in index_symbols or raw_symbol in index_symbols:
            result['security_type'] = 'FUTIDX' if result['segment'] == 'FUTURES' else 'OPTIDX'
        else:
            result['security_type'] = 'FUTSTK' if result['segment'] == 'FUTURES' else 'OPTSTK'

        # Parse expiry date: DD-MMM-YYYY format
        # For FUT: FUT-SYMBOL-DD-MMM-YYYY
        # For OPT: OPT-SYMBOL-DD-MMM-YYYY-CE/PE-STRIKE
        try:
            if len(parts) >= 5:
                # parts[2] = day, parts[3] = month, parts[4] = year
                expiry_str = f"{parts[2]}-{parts[3]}-{parts[4]}"
                result['expiry_date'] = datetime.strptime(expiry_str, '%d-%b-%Y').date()

                # For options, parts[5] = CE/PE, parts[6] = strike
                if contract_type == 'OPT' and len(parts) >= 7:
                    result['option_type'] = parts[5].upper()
                    try:
                        result['strike_price'] = Decimal(parts[6])
                    except (InvalidOperation, ValueError):
                        pass
        except ValueError as e:
            logger.warning(f"Could not parse date from contract: {contract_str}, error: {e}")

        return result

    def import_file(self, file_obj, user=None) -> Dict:
        """
        Import F&O data from Breeze FNO Portfolio CSV file.

        Args:
            file_obj: File object or file path
            user: User performing the import

        Returns:
            dict with import results
        """
        self.batch_id = f"BREEZE_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        self.errors = []
        self.records_created = 0
        self.records_updated = 0
        self.records_skipped = 0
        self.records_unchanged = 0

        # Get filename
        if hasattr(file_obj, 'name'):
            filename = file_obj.name
        else:
            filename = 'breeze_fno.csv'

        # Determine FY from current date (Breeze portfolio shows current FY data)
        self.fy = self._determine_fy_from_current_date()

        # Create import log
        import_log = CSVImportLog.objects.create(
            import_batch_id=self.batch_id,
            file_type='BREEZE_FNO',
            original_filename=filename,
            status='PROCESSING',
            imported_by=user,
        )

        try:
            # Read CSV content
            if hasattr(file_obj, 'read'):
                content = file_obj.read()
                if isinstance(content, bytes):
                    content = content.decode('utf-8-sig')  # Handle BOM
            else:
                with open(file_obj, 'r', encoding='utf-8-sig') as f:
                    content = f.read()

            # Parse CSV
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)

            logger.info(f"Read {len(rows)} rows from Breeze CSV")

            # Find header row (contains "Contract" column)
            header_row_idx = -1
            for i, row in enumerate(rows):
                if row and len(row) > 0:
                    first_cell = str(row[0]).strip().lower()
                    if first_cell == 'contract':
                        header_row_idx = i
                        logger.info(f"Found header row at index {i}")
                        break

            if header_row_idx == -1:
                raise ValueError("Could not find header row with 'Contract' column. Make sure this is a Breeze FNO Portfolio CSV.")

            # Process data rows
            with transaction.atomic():
                for i in range(header_row_idx + 1, len(rows)):
                    row = rows[i]

                    # Skip empty rows
                    if not row or len(row) < 3:
                        continue

                    contract = str(row[0]).strip() if row[0] else ''
                    if not contract:
                        continue

                    # Only process FUT and OPT contracts
                    if not contract.upper().startswith(('FUT-', 'OPT-')):
                        self.records_skipped += 1
                        continue

                    try:
                        self._process_breeze_row(row)
                    except Exception as e:
                        error_msg = f"Row {i+1}: {str(e)}"
                        self.errors.append(error_msg)
                        logger.warning(error_msg)

            # Update import log
            import_log.records_created = self.records_created
            import_log.records_updated = self.records_updated
            import_log.records_skipped = self.records_skipped
            import_log.errors = self.errors
            import_log.status = 'SUCCESS' if not self.errors else 'PARTIAL'
            import_log.save()

            logger.info(f"Import complete: {self.records_created} created, {self.records_updated} updated, {self.records_unchanged} unchanged")

            return {
                'success': True,
                'batch_id': self.batch_id,
                'records_created': self.records_created,
                'records_updated': self.records_updated,
                'records_skipped': self.records_skipped,
                'records_unchanged': self.records_unchanged,
                'errors': self.errors,
            }

        except Exception as e:
            logger.exception(f"Error importing Breeze CSV: {e}")
            import_log.status = 'FAILED'
            import_log.errors = [str(e)]
            import_log.save()

            return {
                'success': False,
                'batch_id': self.batch_id,
                'error': str(e),
                'errors': [str(e)],
            }

    def _process_breeze_row(self, row: List[str]):
        """
        Process a single row from Breeze CSV with incremental upload support.

        Incremental behavior:
        - Creates new records if they don't exist
        - Updates only if values have changed (latest report takes priority)
        - Skips unchanged records to avoid unnecessary database writes

        Column mapping (0-indexed):
        0: Contract
        1: Trade Flow
        2: Open Position Qty
        3: Open Position Value
        4: Open Position Avg. Price
        5: LTP
        6: Profit/Loss Unrealized
        7: Profit/Loss Realized
        """
        contract_str = str(row[0]).strip()

        # Parse contract details
        contract = self.parse_contract(contract_str)

        # Parse numeric values
        open_qty = parse_int(row[2] if len(row) > 2 else '')
        open_value = parse_decimal(row[3] if len(row) > 3 else '')
        unrealized_pnl = parse_decimal(row[6] if len(row) > 6 else '')
        realized_pnl = parse_decimal(row[7] if len(row) > 7 else '')

        # Calculate gross/net P&L (sum of realized + unrealized)
        gross_pnl = realized_pnl + unrealized_pnl
        net_pnl = gross_pnl  # Breeze doesn't show charges in this CSV

        # Determine FY for this record
        # For F&O, ALWAYS use expiry_date to determine FY (when the P&L is realized)
        # This ensures contracts are assigned to the correct FY based on settlement date
        if contract['expiry_date']:
            record_fy = determine_fy_from_date(contract['expiry_date'])
            logger.debug(f"F&O record: using expiry_date {contract['expiry_date']} -> FY {record_fy} for {contract_str}")
        elif self.fy:
            # Fallback to current date FY
            record_fy = self.fy
            logger.debug(f"Using current date FY {record_fy} for {contract_str}")
        else:
            # Last resort: use current date FY
            record_fy = determine_fy_from_date(datetime.now().date())
            logger.debug(f"Using current date FY {record_fy} for {contract_str}")

        # New values from the CSV (latest report)
        new_values = {
            'import_batch_id': self.batch_id,
            'symbol': contract['symbol'],
            'segment': contract['segment'],
            'security_type': contract['security_type'],
            'expiry_date': contract['expiry_date'],
            'strike_price': contract['strike_price'],
            'option_type': contract['option_type'],
            'quantity': open_qty,
            'buy_amount': open_value,
            'sell_amount': Decimal('0.00'),
            'gross_pnl': gross_pnl,
            'net_pnl': net_pnl,
            'realized_pnl': realized_pnl,
            'unrealized_pnl': unrealized_pnl,
            'raw_data': {'row': row},
        }

        # Check if record already exists
        try:
            existing = BrokerContractPnL.objects.get(
                broker='ICICI',
                trading_symbol=contract_str,
                fy=record_fy
            )

            # Compare key fields to detect if anything has changed
            has_changes = False
            for field in self.COMPARE_FIELDS:
                old_value = getattr(existing, field)
                new_value = new_values.get(field)
                if old_value != new_value:
                    has_changes = True
                    logger.debug(f"Change detected in {contract_str}.{field}: {old_value} -> {new_value}")
                    break

            if has_changes:
                # Update with new values (latest report takes priority)
                for key, value in new_values.items():
                    setattr(existing, key, value)
                existing.save()
                self.records_updated += 1
                logger.debug(f"Updated record for {contract_str}: Realized={realized_pnl}, Unrealized={unrealized_pnl}")
            else:
                # No changes detected, skip this record
                self.records_unchanged += 1
                logger.debug(f"Unchanged record for {contract_str}: Realized={realized_pnl}, Unrealized={unrealized_pnl}")

        except BrokerContractPnL.DoesNotExist:
            # Create new record
            BrokerContractPnL.objects.create(
                broker='ICICI',
                trading_symbol=contract_str,
                fy=record_fy,
                **new_values
            )
            self.records_created += 1
            logger.debug(f"Created record for {contract_str}: Realized={realized_pnl}, Unrealized={unrealized_pnl}, FY = {record_fy}")


def delete_import_batch(batch_id: str) -> Dict:
    """
    Delete all records associated with an import batch.

    Args:
        batch_id: The import batch ID to delete

    Returns:
        dict with deletion results
    """
    try:
        with transaction.atomic():
            # Delete contract P&L records
            deleted_pnl = BrokerContractPnL.objects.filter(import_batch_id=batch_id).delete()[0]

            # Delete import log
            deleted_log = CSVImportLog.objects.filter(import_batch_id=batch_id).delete()[0]

            return {
                'success': True,
                'batch_id': batch_id,
                'records_deleted': deleted_pnl,
                'logs_deleted': deleted_log,
            }

    except Exception as e:
        logger.exception(f"Error deleting batch {batch_id}: {e}")
        return {
            'success': False,
            'batch_id': batch_id,
            'error': str(e),
        }
